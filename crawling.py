import io
import sys
import time
import pandas as pd
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# https://ja.dict.naver.com/#/jlpt/list?level=1&part=noun&page=1
# level 1 ~ 5
# page 1 ~ 끝까지 
# allClass = 전체, noun = 명사, verb = 대명사, adjective = 동사, adverb = 조사, 
# measurement = 형용사, phrases = 접사, conjunction = 부사, numeral = 감동사, pronoun = 형용동사, interjection = 기타
# amount는 해당 class의 전체 단어의 수
# page는 .amount / 10 으로 해야할 듯

driver = webdriver.Chrome()
clist = ['noun', 'verb', 'adjective', 'adverb', 'measurement', 'phrases', 'conjunction', 'numeral', 'pronoun', 'interjection']

filename = "japanese_crawling.xlsx"
data = []
                
def append_to_excel(filename, data):
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.create_sheet(index=0, title="Sheet1")
    sheet = wb["Sheet1"]
    
    sheet.append(data)
    
    wb.save(filename)

for level in range(1, 6):
    for word_class in clist:
        url = f'https://ja.dict.naver.com/#/jlpt/list?level={level}&part={word_class}&page=1'
        driver.get(url)
        time.sleep(1)
        amount = driver.find_element(By.CLASS_NAME, 'amount').text
        page_num = int(int(amount.replace(',','').replace('건','')) / 10 + 2)

        for page in range(1, page_num):
            url = f'https://ja.dict.naver.com/#/jlpt/list?level={level}&part={word_class}&page={page}'
            driver.get(url)
            time.sleep(1)

            words = driver.find_elements(By.CLASS_NAME, 'row')
            for word in words:
                chinese = ''
                japanese = ''
                korean = ''
                
                row = word.text.split('\n') # \n 을 기준으로 3개로 나눔.
                
                # row[0]: 'あつりょく [圧力]' --> [이 있으면 [] 안의 문자는 한자, [의 이전 문자들은 히라가나. [이 없으면 전부다 히라가나, 한자는 x
                # row[1]: '듣기' --> 사용 안함
                # row[2]: '명사 압력' --> 명사 이후의 문자열 그대로 뜻.
                
                
                if '[' in row[0]:
                    parts = row[0].split('[')
                    japanese = parts[0].strip()
                    chinese = parts[1].replace(']','').strip()
                else:
                    japanese = row[0].strip()
                korean = row[2][3:]
                
                # data.append([chinese, japanese, korean])
                print([chinese, japanese, korean])
                append_to_excel(filename, [chinese, japanese, korean])

